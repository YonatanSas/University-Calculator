# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Alignment
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# *******************************************FUNCTIONS*****************************************************************

# This function will calculate the bonus
def addBonus(course, units):
    if course in bonuses_dict:
        if units == 4:
            return bonuses_dict[course][0]
        if units == 5:
            return bonuses_dict[course][1]
    return 0


# *********************************************************************************************************************
# This function checks whether the student has all the mandatory courses required to receive a grade and additionally
# returns the missing courses if any
def has_all_needed_courses(student_courses, needed_courses):
    missing = [course for course in needed_courses if course not in student_courses]
    return not missing, missing


# *********************************************************************************************************************
# This function checks whether the student has a minimum of 20 study units so that his grade can be calculated
def has_min_study_units(student_courses_data):
    total_units = sum([course_data[0] for course_data in student_courses_data.values()])
    return total_units >= 20


# *********************************************************************************************************************
# This function calculates the total number of study units the student studied
def totalUnits(student_courses_data):
    return sum([course_data[0] for course_data in student_courses_data.values()])


# *********************************************************************************************************************
# This function calculates the student's final grade according to the following conditions:
# 1) First, calculate the grade for the compulsory courses.
# 2) Switch to another subject and check whether it raises the final grade, if so add it,
#    otherwise it is not considered in the grade
def calculate_final_grade(student_courses_data):
    # Calculate final grade for compulsory courses
    total_required_units = sum(
        [course_data[0] for course, course_data in student_courses_data.items() if course in needed_courses])
    total_required_grade = sum(
        [course_data[0] * course_data[1] for course, course_data in student_courses_data.items() if
         course in needed_courses])

    required_average = total_required_grade / total_required_units

    finalGrade = required_average

    # Iterate over additional courses
    for course, course_data in student_courses_data.items():
        if course not in needed_courses:
            temp_total_units = total_required_units + course_data[0]
            temp_total_grade = total_required_grade + (course_data[0] * course_data[1])
            temp_average = temp_total_grade / temp_total_units

            if temp_average > finalGrade:
                finalGrade = temp_average
                total_required_units = temp_total_units
                total_required_grade = temp_total_grade

    return finalGrade


# *********************************************************************************************************************
def on_exit():
    root.destroy()
    exit()


# *********************************************************************************************************************
def on_cancel():
    root.destroy()
    exit()


# *********************************************************************************************************************
def on_ok():
    global file_path
    file_path = filedialog.askopenfilename()
    if not file_path:
        exit()
    root.destroy()


# *********************************************************************************************************************
def center_window(root, width, height):
    # Get screen width and height
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calculate position
    x = (screen_width / 2) - (width / 2)
    y = (screen_height / 2) - (height / 2)

    root.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
# *********************************************************************************************************************

if __name__ == '__main__':
    # These are the subjects in which you receive a bonus. The first element in this array if the student studied 4
    # units of study in the same subject and the second element in this array if the student studied 5 units of study
    # in the same subject
    bonuses_dict = {}
    extraBonuses_dict = {"תנך": [15, 25], "ספרות": [15, 25], "אנגלית": [15, 25], "הסטוריה": [15, 25],
                         "אזרחות": [15, 25], "מתמטיקה": [15, 35], "פיזיקה": [15, 25], "כימיה": [15, 25],
                         "מדעי המחשב": [15, 25], "ביולוגיה": [15, 25], "מחשבת ישראל": [15, 25]}
    basicBonuses_string = """אופטיקה יישומית,אלקטרוניקה,הבעה עברית, אלקטרוניקה ומחשבים, אלקטרוניקה ותקשורת, אלקטרואופטיקה,
    אמנות, אמנות חזותית, אמנות המחול, אמנות הקולנוע, תולדות האמנות, אמנות שימושית, תיאטרון, מדעי החיים 
    והחקלאות, מיקרוביולוגיה, ביולוגיה חקלאית, ביוטכנולוגיה,בקרה במכונות, בקרת תהליכים, מערכות בקרת תהליכים,גאוגרפיה,
    גאולוגיה,חוק ומשפט,חינוך גופני,חקלאות,חשבונאות,טכנולוגיה מוכללת, מדעי ההנדסה, מדעי הטכנולוגיה, 
    טכנולוגיות הבניה,טלקומוניקציה, בזק, מערכות בזק, מערכות תקשורת ומיתוג,יהדות,יישומי ביוטכנולוגיה,כימיה טכנולוגית, 
    תעשיה כימית,כלכלה ומינהל, מינהל וכלכלה,לימודי ארץ ישראל,לימודי הסביבה, מדעי הסביבה,מדעי ומערכות הבריאות, 
    מערכות רפואיות,מדעי החברה,מדעי החיים, בינה מלאכותית ומערכות מומחה, גרפיקה ממוחשבת, מחשבים ומערכות, 
    מחשוב ובקרה, יסודות תורת המחשב, מדעי ההנדסה, מדע חישובי,מוסיקה, מוסיקה-רסיטל,מחול, מחול- רסיטל,
    מטאורולוגיה,מיכון חקלאי ומערכות, מכונות חקלאיות וטרקטורים, מכטרוניקה,מכניקה הנדסית, פרקי מכונות וחוזק חומרים, 
    תכנון מכונות, מערכות ביוטכנולוגיות,מערכות חשמל, תורת החשמל, מערכות אלקטרואופטיות, מערכות אלקטרוניות,מערכות פיקוד 
    ובקרה, מכשור בקרה ומחשבים, מכשירים ופיקוד, מערכות מכשור ובקרה,מערכות תיכון וייצור, מערכות סיב"מ, מערכות תיב"מ, 
    מערכות מידע וידע,מערכות תעופה, המטוס ומערכותיו, המסוק ומערכותיו,ניהול, ניהול הייצור, ניהול התפעול,ניהול משאבי 
    אנוש,ניתוח מערכות וארגון קבצים, תכנון ותכנות מערכות, מערכות תכנה וחמרה, ענ"א,עברית (לערבים ולדרוזים),
    עולם הערבים והאיסלאם, תולדות הערבים,עם ועולם,ערבית (כשפה זרה), ערבית לערבים,פילוסופיה,שפה זרה,קירור ומיזוג 
    אויר,תיאטרון, ספרות התיאטרון, תולדות התיאטרון,תושב"ע, תלמוד,תיירות, תיירות ותפעול,תכנון יישומים מנהליים, תכנון הנדסי 
    של מבנים, חישוב סטטי וקונסטרוקציות, טכנולוגיות הבנייה,תקשורת, תקשורת המונים, תקשורת וחברה,תרמודינמיקה טכנית, 
    מכונות חום ותרמודינמיקה, מנועי מטוסים ותרמודינמיקה"""

    # Splitting by newline, then by comma, and stripping whitespace
    basicBonuses_list = [profession.strip() for line in basicBonuses_string.split('\n') for profession in
                         line.split(',')]

    # Creating dictionary where each profession maps to [10, 20]
    basicBonuses_dict = {course: [10, 20] for course in basicBonuses_list}

    bonuses_dict.update(extraBonuses_dict)
    bonuses_dict.update(basicBonuses_dict)

    # These are the mandatory courses that the student is required to complete so that his grade can be calculated
    needed_courses_for_grade = ["אנגלית", "מתמטיקה", "הסטוריה", "אזרחות", "הבעה עברית", "ספרות", "תנך"]

    # These are the compulsory courses that will be included in the calculation
    needed_courses = ["אנגלית", "מתמטיקה", "הסטוריה", "אזרחות", "הבעה עברית"]


# *******************************************GUI*****************************************************************

    # Color scheme
    BG_COLOR = "#F0F0F0"
    PRIMARY_COLOR = "#2196F3"
    SECONDARY_COLOR = "#FFFFFF"
    FONT_COLOR = "#333333"
    BUTTON_COLOR = "#1976D2"
    BUTTON_FONT_COLOR = "#FFFFFF"

    # Create a root window for the custom dialog
    root = tk.Tk()
    root.title("The Hebrew University of Jerusalem")
    root.configure(bg=BG_COLOR)

    # Center the window on the screen
    root.eval('tk::PlaceWindow . center')

    # Create a frame for the content
    content_frame = ttk.Frame(root, padding=20)
    content_frame.pack(expand=True, fill=tk.BOTH)

    # Create a label for the title
    title_label = ttk.Label(content_frame, text="University Admission Predictor", font=("Arial", 20, "bold"),
                            foreground=PRIMARY_COLOR)
    title_label.pack(pady=(0, 20))

    # Create a label for the instructions
    instructions_label = ttk.Label(content_frame, text="Please choose the Excel file to proceed.", font=("Arial", 14),
                                   foreground=FONT_COLOR)
    instructions_label.pack()

    # Create a frame for the course names
    course_names_frame = ttk.Frame(content_frame)
    course_names_frame.pack(pady=(5, 10))

    # Create a frame to hold the buttons
    button_frame = ttk.Frame(content_frame)
    button_frame.pack()

    # Create the "Choose File" button
    choose_file_button = tk.Button(button_frame, text="Choose File", command=on_ok, bg=BUTTON_COLOR,
                                   fg=BUTTON_FONT_COLOR, font=("Arial", 12, "bold"), padx=20, pady=10, relief=tk.RAISED,
                                   bd=0, activebackground=PRIMARY_COLOR, activeforeground=SECONDARY_COLOR)
    choose_file_button.pack(side=tk.LEFT, padx=10)

    # Create the "Cancel" button
    cancel_button = tk.Button(button_frame, text="Cancel", command=on_cancel, bg=BUTTON_COLOR, fg=BUTTON_FONT_COLOR,
                              font=("Arial", 12, "bold"), padx=20, pady=10, relief=tk.RAISED, bd=0,
                              activebackground=PRIMARY_COLOR, activeforeground=SECONDARY_COLOR)
    cancel_button.pack(side=tk.LEFT)

    # Update the window size to fit the content
    root.update()
    root.geometry(f"{root.winfo_width()}x{root.winfo_height()}")

    file_path = ""
    root.mainloop()

    if not file_path:
        exit()

    # Define a default file name
    default_file_name = "new_grades.xlsx"

    messagebox.showinfo("Select Save Path", "Please select a path to save the file.")
    # Prompt the user to select a save location
    save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                  title="Choose a location to save the new file",
                                                  initialfile=default_file_name)

    if not save_file_path:
        # Get the user's desktop directory
        messagebox.showerror("Error", "No location was selected to save the file.\nExiting...")
        exit()

    # Use the selected path for saving the Excel file
    output_file_path = save_file_path

    # *********************************************************************************************************************
    # Read the Excel file with the chosen path
    # Read Excel file and select only relevant columns
    columns = ["ת.ז", "מקצוע", "יחידות", "ציון סופי/ מנובא", "האם בוצע ניבוי"]
    df = pd.read_excel(file_path, usecols=columns, engine='openpyxl')

    # Group by student ID and aggregate course data
    student_data = {}
    for index, row in df.iterrows():
        student_id = row["ת.ז"]
        student_course = row["מקצוע"]
        study_units = int(row["יחידות"])
        student_grade = int(row["ציון סופי/ מנובא"])
        final_score_status = row["האם בוצע ניבוי"]

        if student_id not in student_data:
            student_data[student_id] = {}

        # If the student has the same subject twice and the current row is the final score, update the record
        if student_course in student_data[student_id] and final_score_status == "לא בוצע ניבוי, זהו הציון הסופי":
            student_data[student_id][student_course] = [study_units, student_grade]
        elif student_course not in student_data[student_id]:
            # If the student doesn't already have the subject, add it
            student_data[student_id][student_course] = [study_units, student_grade]

    # Apply Bonuses
    for student_id in student_data:
        for student_course in student_data[student_id]:
            study_units, student_grade = student_data[student_id][student_course]
            student_data[student_id][student_course][1] += addBonus(student_course, study_units)

    # Prepare output DataFrame
    output_data = []
    for student_id, student_courses in student_data.items():
        total_units = totalUnits(student_courses)
        has_all, missing = has_all_needed_courses(student_courses, needed_courses_for_grade)
        if has_all and has_min_study_units(student_courses):
            final_grade = calculate_final_grade(student_courses)
            output_data.append([student_id, total_units, "-", final_grade])
        elif has_all:
            output_data.append([student_id, total_units, "-", "לסטודנט יש פחות מ20 יחידות לימוד"])
        else:
            missing_str = ", ".join(missing)
            output_data.append([student_id, total_units, missing_str, "חסר מקצועות חובה"])

    output_df = pd.DataFrame(output_data,
                             columns=["תעודת זהות", "מס' יחידות לימוד", "מקצועות חובה שחסרים", "ציון סופי"])

# ********************************************Create an Excel file********************************************
    writer = pd.ExcelWriter(output_file_path, engine='openpyxl')

    # Write the DataFrame to Excel
    output_df.to_excel(writer, index=False, sheet_name='ציונים סופיים')

    # Get the workbook and the sheet
    workbook = writer.book
    worksheet = writer.sheets['ציונים סופיים']

    # Apply formatting to each cell
    # Define a border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Define a fill for the header
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Define a fill for alternating rows
    row_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_fill_2 = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Apply formatting, border, and fill to each cell
    for row_idx, row in enumerate(worksheet.iter_rows()):
        fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
        # Apply header fill for the first row
        if row_idx == 0:
            fill = header_fill
        for cell in row:
            # Center align the text
            cell.alignment = Alignment(horizontal='center')
            # Apply the border
            cell.border = thin_border
            # Apply the fill
            cell.fill = fill

    # Save the workbook (not the writer)
    workbook.save(output_file_path)  # Saving the file with the user-selected path
    messagebox.showinfo("Success", "File saved successfully!")
